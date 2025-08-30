from openpyxl import load_workbook
from math import gcd
import random
from typing import Optional

def _coprime_a(rng, N):
    while True:
        a = rng.randrange(1, N)      # 1..N-1
        if gcd(a, N) == 1:
            return a

def stream_hyperparams_xlsx(path: str, sheet: str = "grid", seed: Optional[int] = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    header = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    first_data_row = 2

    N = 0
    for row in ws.iter_rows(min_row=first_data_row, values_only=True):
        if all(v is None or (isinstance(v, str) and not str(v).strip()) for v in row):
            break
        N += 1
    if N <= 0:
        wb.close()
        return

    rng = random.Random() if seed is None else random.Random(seed)
    a = _coprime_a(rng, N)
    b = rng.randrange(0, N)

    def read_row(idx1):
        vals = [ws.cell(row=idx1, column=j+1).value for j in range(len(header))]
        return {header[j]: (vals[j] if j < len(vals) else None) for j in range(len(header))}

    for k in range(N):
        i = (a * k + b) % N
        row_idx = first_data_row + i
        yield read_row(row_idx)

    wb.close()
